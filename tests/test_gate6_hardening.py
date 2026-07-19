"""Regression tests for the GATE 6 red-team findings (3 HIGH, 4 MEDIUM)."""

from __future__ import annotations

import warnings
from pathlib import Path

import pytest
from harness.stdf_synth import sample_conditioned_run

from docusearch import analytics, ingest, report_export
from docusearch import config as cfg
from docusearch.server import Service
from docusearch.store import Store


def _spec(body: str) -> dict:  # type: ignore[type-arg]
    return {"title": "R", "sections": [{"heading": "s", "body": body}], "evidence": set()}


# H1 — qq with <2 series raises a clean ValueError, never an IndexError/500.
def test_h1_qq_needs_two_series() -> None:
    with pytest.raises(ValueError, match="two"):
        analytics.render_plot("qq", y=[1, 2, 3])
    with pytest.raises(ValueError, match="two"):
        analytics.render_plot("qq", series=[("only", [1, 2])])


# H2 — an analytics tool on a NON-STDF doc in a mixed store fails clean (ValueError), no pystdf crash.
def test_h2_non_stdf_doc_fails_clean(tmp_path: Path) -> None:
    root = tmp_path / "d"
    root.mkdir()
    (root / "a.html").write_text("<body><h1>Hi</h1><p>plain prose here</p></body>", encoding="utf-8")
    p = tmp_path / "c.yaml"
    p.write_text(
        f'paths:\n  staging_dir: "{(tmp_path / "s").as_posix()}"\n'
        f'  db_path: "{(tmp_path / "c.db").as_posix()}"\n  tmp_dir: "{(tmp_path / "t").as_posix()}"\n'
        f'sources:\n  - name: d\n    location: "{root.as_posix()}"\n    min_content_chars: 1\n'
        'embed:\n  model: "none"\n',
        encoding="utf-8",
    )
    config = cfg.load(p)
    with Store.open(config.paths.db_path) as store:
        ingest.run_ingest(config, store)
        doc_id = next(iter(store.document_path_to_id().values()))
    with pytest.raises(ValueError, match="not a readable STDF"):
        Service(config).stdf_plot(doc_id, 1000)


# H3 — a spreadsheet cell starting with '=' is stored as TEXT (formula injection neutralized).
def test_h3_xlsx_formula_injection_neutralized(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    data = report_export.export_report(fmt="xlsx", **_spec("=1+2 and =cmd|'/c calc'!A1"))  # type: ignore[arg-type]
    p = tmp_path / "r.xlsx"
    p.write_bytes(data)
    wb = load_workbook(p)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            assert cell.data_type != "f"  # never a live formula
    assert report_export.xlsx_cell("=danger()") == "'=danger()"
    assert report_export.xlsx_cell("safe") == "safe"


# M1 — a NUL / control char in the body doesn't crash docx/xlsx export.
@pytest.mark.parametrize("fmt", ["docx", "xlsx", "pdf", "pptx"])
def test_m1_control_chars_do_not_crash_export(tmp_path: Path, fmt: str) -> None:
    data = report_export.export_report(fmt=fmt, **_spec("bad \x00 \x07 char here"))  # type: ignore[arg-type]
    assert isinstance(data, bytes) and len(data) > 300
    assert report_export.xml_safe("a\x00b\x07c") == "abc"


# M2 — stdf.granularity=part emits per-part rollup chunks.
def test_m2_granularity_part_emits_part_chunks(tmp_path: Path) -> None:
    root = tmp_path / "d"
    root.mkdir()
    sample_conditioned_run(root / "r.stdf")
    p = tmp_path / "c.yaml"
    p.write_text(
        f'paths:\n  staging_dir: "{(tmp_path / "s").as_posix()}"\n'
        f'  db_path: "{(tmp_path / "c.db").as_posix()}"\n  tmp_dir: "{(tmp_path / "t").as_posix()}"\n'
        f'sources:\n  - name: d\n    location: "{root.as_posix()}"\n    include: ["*.stdf"]\n'
        '    min_content_chars: 1\n    insertion: "WS1"\n'
        'embed:\n  model: "none"\nstdf:\n  granularity: "part"\n',
        encoding="utf-8",
    )
    config = cfg.load(p)
    with Store.open(config.paths.db_path) as store, warnings.catch_warnings():
        warnings.simplefilter("ignore")
        ingest.run_ingest(config, store)
        parts = store._conn.execute("SELECT COUNT(*) FROM chunks WHERE kind='part'").fetchone()[0]  # noqa: SLF001
    assert parts == 2  # two parts in the sample -> two rollup chunks


# M3 — store_type is actually consumed: health() surfaces it so a client can route.
def test_m3_store_type_consumed_in_health(tmp_path: Path) -> None:
    p = tmp_path / "c.yaml"
    p.write_text(
        f'store_type: "data"\npaths:\n  staging_dir: "{(tmp_path / "s").as_posix()}"\n'
        f'  db_path: "{(tmp_path / "c.db").as_posix()}"\n  tmp_dir: "{(tmp_path / "t").as_posix()}"\n'
        f'sources:\n  - name: d\n    location: "{(tmp_path / "d").as_posix()}"\n'
        'embed:\n  model: "none"\n',
        encoding="utf-8",
    )
    (tmp_path / "d").mkdir()
    health = Service(cfg.load(p)).health()
    assert health["store_type"] == "data" and "stdf_results" in health


# M4 — the plotly backend renders byte-identical for identical inputs (deterministic div id).
def test_m4_plotly_is_deterministic() -> None:
    a = analytics.render_plot("linear", x=[1, 2, 3], y=[1, 4, 9], title="t", backend="plotly")
    b = analytics.render_plot("linear", x=[1, 2, 3], y=[1, 4, 9], title="t", backend="plotly")
    assert a == b


# ---- Phase 6b red-team findings (sessions 18–21 surface) --------------------------------------


def _data_cfg(tmp_path: Path) -> cfg.Config:
    src = tmp_path / "seed"
    src.mkdir()
    p = tmp_path / "d.yaml"
    p.write_text(
        f'store_type: "data"\npaths:\n  staging_dir: "{(tmp_path / "s").as_posix()}"\n'
        f'  db_path: "{(tmp_path / "c.db").as_posix()}"\n  tmp_dir: "{(tmp_path / "t").as_posix()}"\n'
        f'sources:\n  - name: seed\n    location: "{src.as_posix()}"\n'
        '    include: ["*.stdf"]\n    min_content_chars: 1\n    insertion: "FT"\nembed:\n  model: "none"\n',
        encoding="utf-8")
    return cfg.load(p)


def _targz(name_to_bytes: dict) -> bytes:  # type: ignore[type-arg]
    import io
    import tarfile
    buf = io.BytesIO()
    with tarfile.open(fileobj=buf, mode="w:gz") as tar:
        for name, data in name_to_bytes.items():
            info = tarfile.TarInfo(name=name)
            info.size = len(data)
            tar.addfile(info, io.BytesIO(data))
    return buf.getvalue()


def _stdf(vals: list[float]) -> bytes:
    from harness.stdf_synth import StdfBuilder
    b = StdfBuilder().far().mir(lot_id="L", test_cod="WS1")
    for i, v in enumerate(vals):
        b.pir()
        b.ptr(1000, "VMIN", v, lo=0.0, hi=2.0, units="V")
        b.prr(part_id=str(i + 1), hard_bin=1)
    b.mrr()
    return b.to_bytes()


def test_phase6b_h1_upload_sku_traversal_rejected(tmp_path: Path) -> None:
    """H1: a traversal SKU must not escape the store's uploads staging dir."""
    svc = Service(_data_cfg(tmp_path))
    bundle = _targz({"a.stdf": _stdf([0.5, 0.6, 0.7])})
    marker = tmp_path.parent / "sku_escape_marker"
    for bad in ("../../../../sku_escape_marker", "a/b", r"a\b", ".."):
        with pytest.raises(ValueError, match="SKU|label"):
            svc.upload_archive(data=bundle, filename="u.tar.gz", sku=bad, insertion="WS1")
    assert not marker.exists()  # nothing escaped the staging tree
    # a normal SKU still works
    ok = svc.upload_archive(data=bundle, filename="u.tar.gz", sku="WIDGET_9000", insertion="WS1")
    assert ok["documents"] == 1


def test_phase6b_h2_mcp_client_malformed_result_becomes_mcperror() -> None:
    from docusearch.mcp_client import MCPError, _unwrap

    class _Bad:  # a content block claiming text but missing .text
        type = "text"

    class _Result:
        isError = False
        structuredContent = None
        content = [_Bad()]
    # unwrap of a non-compliant result must not raise a raw AttributeError; here it yields "" text
    assert _unwrap(_Result(), "list_stdf") in ("", None)
    # an isError result still raises MCPError (not a raw crash in the error path)
    _Result.isError = True
    with pytest.raises(MCPError):
        _unwrap(_Result(), "list_stdf")


def test_phase6b_h3_nan_results_are_filtered_not_crash() -> None:
    nan = float("nan")
    inf = float("inf")
    assert analytics.summary_stats([1.0, nan, 2.0, inf, 3.0])["n"] == 3  # non-finite dropped
    assert analytics.capability([1.0, nan, 2.0, 3.0, 4.0], 0.0, 5.0)["cpk"] is not None
    # classify no longer mislabels an all-NaN column as "discrete"
    assert analytics.classify_distribution([nan] * 20)["shape"] == "sparse"
    # site_dispersion doesn't return a nan spread / false "agree"
    sd = analytics.site_dispersion({1: [nan, nan, nan], 2: [nan, nan]})
    assert sd["site_shift"] is None and sd["spread_sigma"] is None


def test_phase6b_h3_live_audit_survives_a_nan_result(tmp_path: Path) -> None:
    import struct
    svc = Service(_data_cfg(tmp_path))
    # craft an STDF whose one PTR result is NaN, ingested as a real doc
    raw = bytearray(_stdf([0.5, 0.6, 0.7]))
    nan_bytes = struct.pack("<f", float("nan"))
    good = struct.pack("<f", 0.5)
    raw = bytes(raw).replace(good, nan_bytes, 1)
    store_dir = tmp_path / "in"
    store_dir.mkdir()
    (store_dir / "n.stdf").write_bytes(raw)
    (store_dir / "m.stdf").write_bytes(_stdf([0.55, 0.65, 0.75]))
    import io
    import tarfile
    buf = io.BytesIO()
    with tarfile.open(fileobj=buf, mode="w:gz") as tar:
        tar.add(store_dir / "n.stdf", arcname="n.stdf")
        tar.add(store_dir / "m.stdf", arcname="m.stdf")
    svc.upload_archive(data=buf.getvalue(), filename="u.tar.gz", sku="NANPART", insertion="WS1")
    with Store.open(svc.config.paths.db_path) as st:
        ids = [int(r["id"]) for r in
               st._conn.execute("SELECT id FROM documents WHERE fmt='stdf' ORDER BY id").fetchall()]  # noqa: SLF001
    out = svc.stdf_audit(ids[0], ids[1])  # must not raise
    assert out["html"].startswith("<!doctype html>")


def test_phase6b_h4_stdf_trend_empty_doc_ids(tmp_path: Path) -> None:
    svc = Service(_data_cfg(tmp_path))
    with pytest.raises(ValueError, match="at least one document"):
        svc.stdf_trend([], 1000)


def test_phase6b_h5_renamed_test_number_stays_one_test() -> None:
    from harness.stdf_synth import StdfBuilder

    from docusearch import stdf, stdf_analytics
    b = StdfBuilder().far().mir(lot_id="L", test_cod="WS1")
    b.pir()
    b.ptr(3000, "VMIN_first", 0.5, lo=0.0, hi=2.0, units="V")  # first record names test 3000
    b.prr(part_id="1", hard_bin=1)
    b.pir()
    b.ptr(3000, "VMIN_SECOND", 0.6)  # later touchdown mislabels the SAME test number
    b.prr(part_id="2", hard_bin=1)
    b.mrr()
    run = stdf.parse_stdf_tests(b.to_bytes())
    names = {t.test_txt for t in run.tests if t.test_num == 3000}
    assert names == {"VMIN_first"}  # first-record name is authoritative → not split
    _keys, rows = stdf_analytics.diff_tests(run, run)
    assert len([r for r in rows if r.name == "VMIN_first"]) == 1


def test_phase6b_h6_match_glob_case_insensitive() -> None:
    from docusearch.server import _match_glob
    assert _match_glob("/data/ate/LOTZ_RUN2.stdf", "*run2*")
    assert _match_glob("/data/ate/lotz_run2.stdf", "*RUN2*")
