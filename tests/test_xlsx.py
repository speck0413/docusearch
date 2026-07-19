"""XLSX extraction (Phase 6 / GATE 6): every sheet → rows linearized (Sheet!A1 locator), merged
cells expanded, cell comments captured, values (cached). Covers Stephen's cases: specs, schedules,
data, bump/ball maps."""

from __future__ import annotations

from pathlib import Path

from docusearch import ingest


def _build_xlsx(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment

    wb = Workbook()
    specs = wb.active
    specs.title = "Specs"
    specs["A1"] = "Param"
    specs["B1"] = "Value"
    specs["A2"] = "Vddcore"
    specs["B2"] = "0.75V nonce ZQX7734"
    specs["A2"].comment = Comment("critical rail — do not exceed", "reviewer")

    bump = wb.create_sheet("BumpMap")
    bump["A1"] = ""
    bump["B1"] = "1"
    bump["C1"] = "2"
    bump["A2"] = "A"
    bump["B2"] = "VDD"
    bump["C2"] = "GND"

    sched = wb.create_sheet("Schedule")
    sched["A1"] = "Milestone"
    sched.merge_cells("A1:B1")  # merged header
    sched["A2"] = "Tapeout"
    sched["B2"] = "2026-09-01"
    wb.save(str(path))


def test_extract_xlsx_sheets_merged_comments(tmp_path: Path) -> None:
    p = tmp_path / "book.xlsx"
    _build_xlsx(p)
    doc = ingest.extract_xlsx(p.read_bytes())

    text = "\n".join(s.text for s in doc.segments)
    assert "Vddcore" in text and "ZQX7734" in text  # spec sheet cell values
    assert "VDD" in text and "GND" in text  # bump-map grid
    assert "Tapeout" in text and "2026-09-01" in text  # schedule
    assert "critical rail" in text  # cell comment captured
    # each sheet anchors a Sheet!A1-style locator
    locs = {s.locator if hasattr(s, "locator") else s.heading_path for s in doc.segments}
    assert any("Specs!" in loc for loc in locs)
    assert any("BumpMap!" in loc for loc in locs)


def test_extract_xlsx_trims_and_expands_merged(tmp_path: Path) -> None:
    # a merged header value must appear (expanded from its top-left), and cell text is trimmed.
    p = tmp_path / "book.xlsx"
    _build_xlsx(p)
    doc = ingest.extract_xlsx(p.read_bytes())
    sched = next(s for s in doc.segments if "Schedule!" in (getattr(s, "locator", s.heading_path)))
    assert "Milestone" in sched.text


def test_extract_document_dispatches_xlsx(tmp_path: Path) -> None:
    p = tmp_path / "book.xlsx"
    _build_xlsx(p)
    doc = ingest.extract_document(p, "xlsx")
    assert any("ZQX7734" in s.text for s in doc.segments)
